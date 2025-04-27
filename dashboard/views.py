from django.shortcuts import render
from sign_in.models import SignInRecord, StudentReference
from django.utils.timezone import localtime
from django.utils.dateparse import parse_date
from users.decorators import workstudy_required, admin_required
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.http import HttpResponse
from django.utils import timezone
import pytz
import os
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from django.conf import settings

def localize_to_central(x):
    if isinstance(x, pd.Timestamp):
        if timezone.is_naive(x):
            x = timezone.make_aware(x, timezone.utc)
        return timezone.localtime(x, timezone.get_fixed_timezone(-300))
    return x

def time_format(x):
    if isinstance(x, pd.Timestamp):
        return x.strftime('%I:%M %p')
    return x

@workstudy_required
def data_page(request):
    # Get start date and end date from GET request parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Retrieve the records
    records = SignInRecord.objects.select_related('student')

    # Filter by start date and end date if provided
    if start_date:
        records = records.filter(time_in__gte=start_date)
    if end_date:
        records = records.filter(time_in__lte=end_date)

    # Convert the records to a list of dictionaries, including related fields
    records_list = []
    for record in records:
        # Try/Except loop purges entries with incomplete information or broken student references
        try:
            data = {
                'id': record.id,
                'student_id': record.student.student_id,
                'first_name': record.student.first_name,
                'last_name': record.student.last_name,
                'reason': record.reason,
                'formatted_reason': record.formatted_reason(),
                'time_in': record.time_in,
                'time_out': record.time_out,
                'formatted_date': record.time_in.strftime('%a, %b %d, %Y') if record.time_in else None,
            }
            records_list.append(data)
        except:
            pass

    '''
    records_list = [
        {
            'id': record.id,
            'student_id': record.student.student_id,
            'first_name': record.student.first_name,
            'last_name': record.student.last_name,
            'reason': record.reason,
            'formatted_reason': record.formatted_reason(),
            'time_in': record.time_in,
            'time_out': record.time_out,
            'formatted_date': record.time_in.strftime('%a, %b %d, %Y') if record.time_in else None,
        }
        for record in records
    ]
    '''

    # Convert the list of records to a pandas DataFrame
    df = pd.DataFrame(records_list)

    # Ensure that time_in and time_out are in datetime format
    df['time_in'] = pd.to_datetime(df['time_in'], errors='coerce')
    df['time_out'] = pd.to_datetime(df['time_out'], errors='coerce')

    # Convert time to the local America/Chicago timezone
    df['time_in'] = df['time_in'].apply(localize_to_central)
    df['time_out'] = df['time_out'].apply(localize_to_central)

    # Handle sorting
    sort_by = request.GET.get('sort_by', 'id')
    sort_order = request.GET.get('sort_order', 'desc')
    
    if sort_order == 'desc':
        df = df.sort_values(by=[sort_by], ascending=False)
    else:
        df = df.sort_values(by=[sort_by], ascending=True)
    
    # Fill N/A's
    df['time_in'] = df['time_in'].fillna('No Time In')
    df['time_out'] = df['time_out'].fillna('No Time Out')
    # Applying string formatting to time
    df['time_in'] = df['time_in'].apply(time_format)
    df['time_out'] = df['time_out'].apply(time_format)

    # After sorting, pass cleaned DataFrame to context
    context = {
        'records': df.to_dict('records'),
        'sort_by': sort_by,
        'sort_order': sort_order,
        'start_date': start_date,
        'end_date': end_date,
    }

    return render(request, 'dashboard/data_page.html', context)

def timify(x):
    if isinstance(x, pd.Timestamp):
        return x.time()
    return x

@admin_required
def export_records(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    sort_by = request.GET.get('sort_by', 'id')
    sort_order = request.GET.get('sort_order', 'desc')
    file_date = ""
    records = SignInRecord.objects.all().select_related('student')

    if start_date:
        records = records.filter(time_in__date__gte=parse_date(start_date))
        file_date = " from " + start_date
    if end_date:
        records = records.filter(time_in__date__lte=parse_date(end_date))
        file_date = file_date + " to " + end_date

    records_list = [
        {
            'id': record.id,
            'student_id': record.student.student_id,
            'first_name': record.student.first_name,
            'last_name': record.student.last_name,
            'reason': record.reason,
            'formatted_reason': record.formatted_reason(),
            'time_in': record.time_in if record.time_in else None,
            'time_out': record.time_out if record.time_out else None,
            'formatted_date': record.time_in.date() if record.time_in else None,
        }
        for record in records
    ]

    df = pd.DataFrame(records_list)

    # Same time-formatting logic from above
    df['time_in'] = pd.to_datetime(df['time_in'], errors='coerce')
    df['time_out'] = pd.to_datetime(df['time_out'], errors='coerce')
    df['time_in'] = df['time_in'].fillna('')
    df['time_out'] = df['time_out'].fillna('')
    df['time_in'] = df['time_in'].apply(localize_to_central)
    df['time_out'] = df['time_out'].apply(localize_to_central)

    if sort_order == 'desc':
        df = df.sort_values(by=[sort_by], ascending=False)
    else:
        df = df.sort_values(by=[sort_by], ascending=True)

    df['time_in'] = df['time_in'].apply(timify)
    df['time_out'] = df['time_out'].apply(timify)

    df.drop(['id', 'reason'], axis=1, inplace=True)
    df.columns = ['Student ID', 'First Name', 'Last Name', 'Reason', 'Time In', 'Time Out', 'Date']

    template_path = os.path.join(settings.BASE_DIR, 'static', 'dashboard', 'excel', 'template.xlsx')
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Clearing old data
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # Write DataFrame rows starting from row 2 (row 1 assumed for headers)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Prepare the HTTP response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Library Sign-In Records{file_date}.xlsx"'
    wb.save(response)
    return response